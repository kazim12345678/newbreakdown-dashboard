import streamlit as st
import pandas as pd
import numpy as np
import re
import os
from pathlib import Path
import io
import matplotlib.pyplot as plt
import datetime as dt

from openpyxl import load_workbook  # for saving back to XLSM with keep_vba=True

# ======================================================
# Page setup
# ======================================================
st.set_page_config(page_title="Maintenance KPI Dashboard", layout="wide")
st.title("🛠 Maintenance KPI Dashboard")

# ======================================================
# Persistent storage folder
# ======================================================
DATA_DIR = Path("app_files")        # permanent folder
DATA_DIR.mkdir(parents=True, exist_ok=True)

# ======================================================
# Helper functions
# ======================================================
def list_saved_files():
    return sorted([
        p.name for p in DATA_DIR.iterdir()
        if p.is_file() and p.suffix.lower() in [".xlsx", ".xlsm", ".xls"]
    ])

def save_uploaded_file(uploaded_file, overwrite=False):
    """Save uploaded file bytes to disk permanently.
    This preserves xlsm macros because we do not modify content, only copy bytes. [3](https://stackoverflow.com/questions/76893985/pandas-corrupting-file-when-writing-data-from-xlsx-to-xlsm)
    """
    target = DATA_DIR / uploaded_file.name
    if target.exists() and not overwrite:
        return False, f"File already exists: {uploaded_file.name}"
    with open(target, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return True, f"Saved permanently: {uploaded_file.name}"

def read_excel_smart(file_path_or_buffer):
    """Pick 'Main Data' if present, else first sheet."""
    xls = pd.ExcelFile(file_path_or_buffer)
    sheet = next((s for s in xls.sheet_names if s.strip().lower() in ["main data", "maindata", "main"]), xls.sheet_names[0])
    df = pd.read_excel(file_path_or_buffer, sheet_name=sheet)
    return df, sheet

def to_hours(series):
    """Convert Excel time/duration representations into hours."""
    # timedelta -> hours
    if pd.api.types.is_timedelta64_dtype(series):
        return series.dt.total_seconds() / 3600

    # object -> could be time, datetime, number, string
    def conv(x):
        if pd.isna(x):
            return np.nan
        if isinstance(x, pd.Timedelta):
            return x.total_seconds() / 3600
        if isinstance(x, pd.Timestamp):
            return x.hour + x.minute / 60 + x.second / 3600
        if isinstance(x, dt.time):
            return x.hour + x.minute / 60 + x.second / 3600
        if isinstance(x, dt.datetime):
            return x.hour + x.minute / 60 + x.second / 3600
        if isinstance(x, (int, float, np.integer, np.floating)):
            # Excel fraction-of-day -> hours
            if x <= 1.5:
                return x * 24
            return float(x)
        if isinstance(x, str):
            t = x.strip()
            m = re.match(r"^(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?$", t)
            if m:
                hh = int(m.group(1)); mm = int(m.group(2)); ss = int(m.group(3) or 0)
                return hh + mm / 60 + ss / 3600
        return np.nan

    return series.apply(conv)

def get_hour(x):
    """Extract hour-of-day (0-23) from time/time-like values."""
    if pd.isna(x):
        return np.nan
    if isinstance(x, pd.Timestamp):
        return x.hour
    if isinstance(x, dt.time):
        return x.hour
    if isinstance(x, dt.datetime):
        return x.hour
    if isinstance(x, str):
        m = re.match(r"^(\d{1,2}):", x.strip())
        if m:
            return int(m.group(1))
    return np.nan

def split_names(s):
    """Split technician string 'A/B/C' or 'A & B' etc. into individual names."""
    parts = re.split(r"[/,&]|\band\b", str(s), flags=re.IGNORECASE)
    parts = [p.strip() for p in parts if p.strip() and p.strip().lower() != "nan"]
    return parts if parts else ["Unknown"]

def real_rows_only(df):
    """Remove template/blank rows. A row is 'real' if any key fields exist."""
    keys = ["Notification No.", "Machine No.", "Type", "Reported Problem"]
    present = [c for c in keys if c in df.columns]
    if not present:
        return df.copy()

    mask = False
    for c in present:
        mask = mask | df[c].notna()
    return df[mask].copy()

def df_to_xlsx_bytes(df, sheet_name="Main Data"):
    """Convert dataframe to downloadable xlsx bytes."""
    buff = io.BytesIO()
    with pd.ExcelWriter(buff, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    buff.seek(0)
    return buff.getvalue()

# ======================================================
# Sidebar: File Manager (Permanent)
# ======================================================
st.sidebar.header("📁 File Manager (Permanent storage)")

tab_add, tab_manage, tab_edit = st.sidebar.tabs(["➕ Add", "🗂 Manage", "✏️ Edit"])

with tab_add:
    up = st.file_uploader("Upload Excel (xlsm/xlsx/xls)", type=["xlsm", "xlsx", "xls"], key="upload_save")
    overwrite = st.checkbox("Overwrite if exists", value=False, key="overwrite_save")
    if st.button("Save permanently to folder", use_container_width=True, key="btn_save_file"):
        if up is None:
            st.warning("Please upload a file first.")
        else:
            ok, msg = save_uploaded_file(up, overwrite=overwrite)
            (st.success if ok else st.error)(msg)
            if ok:
                st.rerun()

with tab_manage:
    files = list_saved_files()
    if not files:
        st.info("No saved files yet. Upload one in ➕ Add.")
    else:
        selected_manage = st.selectbox("Select saved file", files, key="manage_select")
        fpath = DATA_DIR / selected_manage

        # Download
        with open(fpath, "rb") as f:
            st.download_button(
                "⬇️ Download selected file",
                data=f.read(),
                file_name=selected_manage,
                use_container_width=True,
                key="btn_download_saved"
            )  # st.download_button [5](https://stackoverflow.com/questions/75528026/saving-files-from-streamlit-into-a-temporary-directory)

        # Rename
        new_name = st.text_input("Rename to (keep extension)", value=selected_manage, key="rename_input")
        if st.button("Rename", use_container_width=True, key="btn_rename"):
            new_path = DATA_DIR / new_name
            if new_path.exists():
                st.error("A file with that name already exists.")
            else:
                fpath.rename(new_path)
                st.success("Renamed.")
                st.rerun()

        # Delete
        if st.button("🗑️ Delete permanently", type="primary", use_container_width=True, key="btn_delete"):
            fpath.unlink(missing_ok=True)
            st.success("Deleted permanently.")
            st.rerun()

with tab_edit:
    files = list_saved_files()
    if not files:
        st.info("No files to edit yet.")
    else:
        selected_edit = st.selectbox("Select file to edit", files, key="edit_select")
        fpath = DATA_DIR / selected_edit

        df_edit, used_sheet = read_excel_smart(fpath)
        st.caption(f"Loaded sheet: {used_sheet} | Rows: {len(df_edit):,} | Columns: {len(df_edit.columns)}")

        # Editable table
        edited_df = st.data_editor(df_edit, num_rows="dynamic", use_container_width=True, key="data_editor")  # [4](https://www.iditect.com/program-example/python--how-to-save-xlsm-file-with-macro-using-openpyxl.html)

        st.markdown("**Save options:**")
        col1, col2 = st.columns(2)
        with col1:
            out_name = st.text_input("Save as (new file name)", value=f"{Path(selected_edit).stem}_edited.xlsx", key="out_name")
        with col2:
            overwrite_out = st.checkbox("Overwrite output if exists", value=False, key="overwrite_out")

        # Save as new xlsx (always)
        if st.button("💾 Save edited as NEW .xlsx", use_container_width=True, key="btn_save_new_xlsx"):
            out_path = DATA_DIR / out_name
            if out_path.exists() and not overwrite_out:
                st.error("Output file exists. Choose another name or enable overwrite.")
            else:
                xbytes = df_to_xlsx_bytes(edited_df, sheet_name="Main Data")
                with open(out_path, "wb") as f:
                    f.write(xbytes)
                st.success(f"Saved: {out_path.name}")
                st.rerun()

        # If original is XLSM, save back to same file preserving macros using keep_vba=True
        if fpath.suffix.lower() == ".xlsm":
            st.info("XLSM detected: You can save back into the SAME XLSM while preserving macros using keep_vba=True. [1](https://cheat-sheet.streamlit.app/)[2](blob:https://fa000000124.resources.office.net/fb93a13e-8828-4905-b110-ad10ba214d90)")
            if st.button("💾 Save back to SAME .xlsm (keep macros)", use_container_width=True, key="btn_save_back_xlsm"):
                wb = load_workbook(fpath, keep_vba=True)  # keep_vba=True preserves VBA project when saving [1](https://cheat-sheet.streamlit.app/)
                # write into 'Main Data' sheet (create if missing)
                ws_name = "Main Data"
                if ws_name in wb.sheetnames:
                    ws = wb[ws_name]
                    # clear existing rows
                    if ws.max_row > 0:
                        ws.delete_rows(1, ws.max_row)
                else:
                    ws = wb.create_sheet(ws_name)

                # header
                for j, col in enumerate(edited_df.columns, start=1):
                    ws.cell(row=1, column=j, value=str(col))

                # data
                for i in range(len(edited_df)):
                    row = edited_df.iloc[i].tolist()
                    for j, val in enumerate(row, start=1):
                        ws.cell(row=i+2, column=j, value=val)

                wb.save(fpath)
                st.success("Saved back to original XLSM with keep_vba=True.")
                st.rerun()

# ======================================================
# MAIN: Choose file to analyze
# ======================================================
st.subheader("📊 Analyze & Dashboard")

mode = st.radio("Choose data source:", ["Use a saved file (permanent)", "Upload once (not saved)"], horizontal=True, key="source_mode")

if mode == "Use a saved file (permanent)":
    files = list_saved_files()
    if not files:
        st.warning("No saved files found. Upload one using the sidebar ➕ Add.")
        st.stop()
    selected = st.selectbox("Select saved file to analyze", files, key="analyze_saved")
    file_to_read = DATA_DIR / selected
else:
    tmp_up = st.file_uploader("Upload file to analyze (not saved)", type=["xlsm", "xlsx", "xls"], key="analyze_once")
    if tmp_up is None:
        st.stop()
    file_to_read = tmp_up

# ======================================================
# Read + clean + compute
# ======================================================
df, sheet = read_excel_smart(file_to_read)
df.columns = [str(c).strip() for c in df.columns]

st.caption(f"Loaded sheet: **{sheet}** | Rows: **{len(df):,}** | Cols: **{len(df.columns)}**")

real = real_rows_only(df)

# Parse Date
if "Date" in real.columns:
    real["Date"] = pd.to_datetime(real["Date"], errors="coerce")

# Convert time columns to hours
if "Time Consumed" in real.columns:
    real["time_h"] = to_hours(real["Time Consumed"])
else:
    real["time_h"] = np.nan

if "Waiting Time" in real.columns:
    real["wait_h"] = to_hours(real["Waiting Time"])
else:
    real["wait_h"] = np.nan

# Hour of day from Start else Requested Time
real["hour"] = np.nan
if "Start" in real.columns:
    real["hour"] = real["Start"].apply(get_hour)
if "Requested Time" in real.columns:
    real.loc[real["hour"].isna(), "hour"] = real.loc[real["hour"].isna(), "Requested Time"].apply(get_hour)

# ======================================================
# Filters
# ======================================================
st.sidebar.header("🔎 KPI Filters")
st.sidebar.write(f"Real rows detected: **{len(real):,}** (from {len(df):,})")

df_f = real.copy()

if "Date" in df_f.columns:
    dmin, dmax = df_f["Date"].min(), df_f["Date"].max()
    if pd.notna(dmin) and pd.notna(dmax):
        start_date, end_date = st.sidebar.date_input(
            "Date range",
            value=(dmin.date(), dmax.date()),
            min_value=dmin.date(),
            max_value=dmax.date(),
            key="date_range"
        )
        df_f = df_f[(df_f["Date"].dt.date >= start_date) & (df_f["Date"].dt.date <= end_date)]

def mfilter(col, label):
    global df_f
    if col in df_f.columns:
        opts = sorted(df_f[col].dropna().astype(str).unique())
        sel = st.sidebar.multiselect(label, opts, key=f"filter_{col}")
        if sel:
            df_f = df_f[df_f[col].astype(str).isin(sel)]

mfilter("Area", "Area")
mfilter("Shift", "Shift")
mfilter("Type", "Type")
mfilter("Machine No.", "Machine No.")
mfilter("Performed By", "Technician")

st.sidebar.caption(f"Filtered rows: **{len(df_f):,}**")

# ======================================================
# KPIs
# ======================================================
total_jobs = len(df_f)
unique_complaints = df_f["Notification No."].nunique(dropna=True) if "Notification No." in df_f.columns else np.nan
total_hours = df_f["time_h"].sum(skipna=True)
avg_hours = df_f["time_h"].mean(skipna=True)

st.subheader("✅ KPI Summary")
c1, c2, c3, c4 = st.columns(4)
c1.metric("Total Jobs (real rows)", f"{total_jobs:,}")
c2.metric("Complaints (unique notifications)", f"{int(unique_complaints):,}" if pd.notna(unique_complaints) else "-")
c3.metric("Total Time Consumed (hours)", f"{total_hours:,.2f}")
c4.metric("Avg Time per Job (hours)", f"{avg_hours:,.2f}" if pd.notna(avg_hours) else "-")

st.divider()

# ======================================================
# Chart 1: Top machines by downtime hours
# ======================================================
st.subheader("🏭 Machine-wise Breakdown (Top 10 by hours)")
if "Machine No." in df_f.columns:
    top_m = df_f.groupby("Machine No.")["time_h"].sum().sort_values(ascending=False).head(10)
    fig, ax = plt.subplots(figsize=(10, 4))
    ax.bar(top_m.index.astype(str), top_m.values)
    ax.set_ylabel("Hours")
    ax.set_xlabel("Machine")
    ax.set_title("Top 10 Machines by Total Time Consumed (hours)")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    st.pyplot(fig)
else:
    st.info("Machine No. column not found.")

# ======================================================
# Chart 2: Technician worked hours (split + equal allocation)
# ======================================================
st.subheader("👷 Technician Worked Hours (Top 10)")
if "Performed By" in df_f.columns:
    rows = []
    for who, h in zip(df_f["Performed By"].fillna("Unknown"), df_f["time_h"]):
        if pd.isna(h):
            continue
        names = split_names(who)
        share = h / len(names)
        for n in names:
            rows.append((n, share))

    tech_df = pd.DataFrame(rows, columns=["Technician", "hours"])
    top_t = tech_df.groupby("Technician")["hours"].sum().sort_values(ascending=False).head(10)

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.bar(top_t.index.astype(str), top_t.values)
    ax.set_ylabel("Hours (allocated)")
    ax.set_xlabel("Technician")
    ax.set_title("Top 10 Technicians by Allocated Worked Hours")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    st.pyplot(fig)
else:
    st.info("Performed By column not found.")

st.divider()

# ======================================================
# Chart 3: Complaints received trend (Date-wise)
# ======================================================
st.subheader("📩 Complaints Received Trend (Date-wise)")
if "Date" in df_f.columns:
    dated = df_f.dropna(subset=["Date"]).copy()
    dated["day"] = dated["Date"].dt.date
    if "Notification No." in dated.columns:
        comp = dated.dropna(subset=["Notification No."]).groupby("day")["Notification No."].nunique()
    else:
        comp = dated.groupby("day").size()

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(comp.index, comp.values)
    ax.set_ylabel("Complaints (unique)")
    ax.set_xlabel("Date")
    ax.set_title("Complaints Received per Day")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    st.pyplot(fig)
else:
    st.info("Date column not found.")

# ======================================================
# Chart 4: 0–23 hour pattern
# ======================================================
st.subheader("🕐 0–23 Hour Pattern (Total Time Consumed)")
hourly = df_f.dropna(subset=["hour"]).groupby("hour")["time_h"].sum().reindex(range(24), fill_value=0)

fig, ax = plt.subplots(figsize=(10, 4))
ax.bar(hourly.index, hourly.values)
ax.set_xlabel("Hour (0–23)")
ax.set_ylabel("Hours")
ax.set_title("Total Time Consumed by Hour of Day")
ax.set_xticks(range(24))
plt.tight_layout()
st.pyplot(fig)

# ======================================================
# Chart 5: Date × Hour heatmap
# ======================================================
st.subheader("🗓️ Date × Hour Heatmap (Time Consumed)")
heat = df_f.dropna(subset=["Date", "hour"]).copy()
heat["day"] = heat["Date"].dt.date

pivot = heat.pivot_table(index="day", columns="hour", values="time_h", aggfunc="sum").reindex(columns=range(24)).fillna(0)
pivot_recent = pivot.tail(30)  # last 30 days in filtered data

fig, ax = plt.subplots(figsize=(12, 5))
im = ax.imshow(pivot_recent.values, aspect="auto", interpolation="nearest")
ax.set_title("Date × Hour Heatmap (last 30 days in filtered data)")
ax.set_xlabel("Hour of day")
ax.set_ylabel("Date")
ax.set_xticks(range(24))
ax.set_xticklabels(range(24))
ax.set_yticks(range(len(pivot_recent.index)))
ax.set_yticklabels([str(d) for d in pivot_recent.index])
fig.colorbar(im, ax=ax, label="Hours")
plt.tight_layout()
st.pyplot(fig)

st.divider()

# ======================================================
# Chart 6: Top 10 breakdown reasons
# ======================================================
st.subheader("🧾 Top 10 Breakdown Reasons")
if "Reported Problem" in df_f.columns:
    reason = df_f["Reported Problem"].fillna("").astype(str).str.strip()
    reason_clean = (reason.str.lower()
                    .str.replace(r"[^a-z0-9\s]", "", regex=True)
                    .str.replace(r"\s+", " ", regex=True)
                    .str.strip())
    reason_short = reason_clean.apply(lambda s: " ".join(s.split()[:6]) if s else "")

    fallback = df_f["Type"].fillna("unknown").astype(str).str.strip().str.lower() if "Type" in df_f.columns else "unknown"
    df_f["reason"] = reason_short.where(reason_short != "", fallback)

    top_r = df_f.groupby("reason")["time_h"].sum().sort_values(ascending=False).head(10)

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.barh(top_r.index[::-1], top_r.values[::-1])
    ax.set_xlabel("Hours")
    ax.set_title("Top 10 Reasons by Total Time Consumed")
    plt.tight_layout()
    st.pyplot(fig)
else:
    st.info("Reported Problem column not found.")

# ======================================================
# Data preview + download filtered data
# ======================================================
st.subheader("📄 Filtered Data")
with st.expander("View filtered table"):
    st.dataframe(df_f, use_container_width=True)

# Download filtered as xlsx
download_bytes = df_to_xlsx_bytes(df_f, sheet_name="Filtered Data")
st.download_button(
    "⬇️ Download filtered data (xlsx)",
    data=download_bytes,
    file_name="filtered_data.xlsx",
    use_container_width=True,
    key="download_filtered"
)  # st.download_button [5](https://stackoverflow.com/questions/75528026/saving-files-from-streamlit-into-a-temporary-directory)
